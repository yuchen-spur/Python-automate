import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb=openpyxl.load_workbook('example.xlsx')
'''
sheet=wb['Sheet1']
'''
anotherSheet=wb.active
'''
print(sheet['A1'].value)
c=sheet['B1']
print(c.value)
print('Row '+str(c.row)+',Column '+str(c.column)+' is '+str(c.value))
print(c.coordinate)#B1

for i in range (1,8,2):
    print(i,sheet.cell(i,2).value)

print('max row:',sheet.max_row,'max column:',sheet.max_column)

print(get_column_letter(27))
print(column_index_from_string('AA'))

for cellofrowobjects in sheet['A1':'C3']:
    for cellobj in cellofrowobjects:
        print(cellobj.coordinate,cellobj.value)
    print('----END OF ROW',cellobj.row,'----')

for cellobj in list(sheet.columns)[1]:
    print(cellobj.value)
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
wb.create_sheet(index=0,title='the first sheet')
wb.create_sheet(index=2,title='the third sheet')
del wb['the third sheet']
sheet=wb['Spam Bacon Eggs Sheet']
sheet['A1']='eggs'

sheet['C9'] = '=SUM(C1:C8)'
'''
anotherSheet.row_dimensions[1].height=70
anotherSheet.column_dimensions['A'].width=20
anotherSheet.merge_cells('A1:D1')

wb.save('example_copy.xlsx')
