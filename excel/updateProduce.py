#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
update_prices={'Celery':1.19,
               'Garlic':3.07,
               'Lemon':1.27,
               }
# Set the font
Fontsize=Font(size=24)
Fontitalic=Font(italic=True)

# TODO: Loop through the rows and update the prices.
for row in range(2,sheet.max_row+1):
    if sheet.cell(row,1).value in update_prices:
        sheet.cell(row,2).value = update_prices[sheet.cell(row,1).value]
        sheet.cell(row,2).font=Fontsize
    if sheet.cell(row,2).value > 5:
        for column in range (1,sheet.max_column+1):
            sheet.cell(row,column).font=Fontitalic

sheet.freeze_panes='A2'        

wb.save('updatedProduceSales.xlsx')
